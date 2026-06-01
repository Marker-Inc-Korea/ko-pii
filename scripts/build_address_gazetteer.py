#!/usr/bin/env python3
"""행정안전부 도로명주소 DB → 건물명 가제티어 빌드.

juso 원본(공동주택/건물 DB)을 증류해 ``building_names.txt.gz`` 로 만든다.
**런타임이 아니라 빌드타임 도구** — 패키지에 넣을 정적 리소스를 한 번 생성한다.

소스 받기:
    business.juso.go.kr → 주소제공 → 공동주택 DB(권장, 가벼움) 또는 건물 DB
    (KOGL Type 1, 회원가입 필요). 압축 해제하면 ``|`` 구분 cp949 텍스트.

사용:
    python scripts/build_address_gazetteer.py \
        --src ~/Downloads/공동주택_전체분.txt \
        --col 9 --encoding cp949 \
        --out src/ko_pii/dictionaries/building_names.txt.gz \
        --max 50000

    # --col 은 건물명/단지명이 들어있는 0-기반 컬럼 인덱스.
    #   파일 첫 줄을 열어 어느 컬럼이 단지명인지 확인 후 지정.

현재 번들된 가제티어 재현 (K-apt Excel, 단지명 6열, 헤더 2행):
    python scripts/build_address_gazetteer.py \
        --src 20260529_단지_기본정보.xlsx --col 5 --header-row 2 \
        --source "K-apt 관리비공개의무단지 기본정보, data.go.kr" \
        --out src/ko_pii/dictionaries/building_names.txt.gz
    # 번들본은 위 결과 + 시드 랜드마크(그랑서울·코엑스 등 K-apt 외 상업건물) 합집합.

접미사(빌딩/타워/자이/래미안 등)로 끝나는 이름은 패턴 룰이 이미 처리하므로
가제티어에서 제외한다 — 비접미사 고유명만 남겨 용량·오탐을 줄인다.
"""
from __future__ import annotations

import argparse
import gzip
import sys
from pathlib import Path

# 패턴 룰이 이미 잡는 접미사 — 가제티어에서 제외 (address._BLDG_SUFFIX 와 동기화)
_RULE_SUFFIXES = (
    "빌딩", "타워", "센터", "스퀘어", "플라자", "프라자", "오피스텔",
    "아파트", "맨션", "하이츠", "캐슬", "팰리스", "레지던스", "펜트하우스",
    "자이", "래미안", "푸르지오", "더샵", "아이파크", "힐스테이트", "디에이치",
    "e편한세상", "위브", "센트레빌", "롯데캐슬", "데시앙", "스위첸", "꿈에그린",
    "베르디움", "리슈빌", "코아루", "우미린", "한라비발디", "효성해링턴",
    "어울림", "하늘채", "호반써밋", "아크로", "써밋", "주공",
)


def _clean(name: str) -> str | None:
    name = name.strip().strip('"').strip()
    # 검출기는 한 토큰만 매칭 → 공백 있으면 마지막 토큰(보통 단지명 핵심)
    if " " in name or "\t" in name:
        name = name.split()[-1]
    if not (4 <= len(name) <= 20):  # 길이 4+ : 짧은 모호어 FP 방지
        return None
    if name.isdigit():
        return None
    if any(name.endswith(sfx) for sfx in _RULE_SUFFIXES):
        return None  # 룰이 이미 처리
    # 한글/영숫자만 (특수문자 포함 이름은 토큰 매칭 불가)
    if not all("가" <= c <= "힣" or c.isalnum() for c in name):
        return None
    # 일반어 제외 (FP 방지) — ko_pii 사전 사용
    try:
        from ko_pii.dictionaries.common_words import is_common_word
        if is_common_word(name):
            return None
    except ImportError:
        pass
    return name


def main() -> int:
    ap = argparse.ArgumentParser(description="juso DB → 건물명 가제티어")
    ap.add_argument("--src", required=True, type=Path, help="juso 원본 텍스트")
    ap.add_argument("--col", required=True, type=int, help="건물명 컬럼 (0-기반)")
    ap.add_argument("--delim", default="|", help="구분자 (기본 |)")
    ap.add_argument("--encoding", default="cp949", help="인코딩 (기본 cp949)")
    ap.add_argument("--out", required=True, type=Path, help="출력 .txt.gz")
    ap.add_argument("--max", type=int, default=0, help="최대 항목 수 (0=무제한)")
    ap.add_argument("--header-row", type=int, default=0,
                    help="헤더 행 번호(1-기반). 데이터는 그 다음부터. 0=헤더 없음")
    ap.add_argument("--source", default="행정안전부 도로명주소 DB (business.juso.go.kr, KOGL Type 1)",
                    help="가제티어 헤더에 표기할 출처")
    args = ap.parse_args()

    if not args.src.exists():
        print(f"[err] 소스 없음: {args.src}", file=sys.stderr)
        return 1

    names: set[str] = set()
    rows = 0
    if args.src.suffix.lower() in (".xlsx", ".xlsm"):
        # K-apt 등 Excel — read_only 는 차원 메타 누락 시 오파싱 → 일반 모드
        import openpyxl  # dev 의존성 (pip install openpyxl)
        ws = openpyxl.load_workbook(args.src, data_only=True).active
        for r in range(args.header_row + 1, ws.max_row + 1):
            rows += 1
            val = ws.cell(row=r, column=args.col + 1).value  # col 0-기반 → 1-기반
            cleaned = _clean(str(val)) if val is not None else None
            if cleaned:
                names.add(cleaned)
    else:
        with args.src.open(encoding=args.encoding, errors="replace") as f:
            for i, line in enumerate(f):
                if i < args.header_row:
                    continue
                rows += 1
                parts = line.rstrip("\n").split(args.delim)
                if args.col >= len(parts):
                    continue
                cleaned = _clean(parts[args.col])
                if cleaned:
                    names.add(cleaned)

    ordered = sorted(names)
    if args.max and len(ordered) > args.max:
        ordered = ordered[: args.max]

    header = [
        "# ko-pii 건물명/단지명 가제티어",
        f"# 출처: {args.source}",
        f"# 원본 {rows:,}행 → 비접미사 고유명 {len(ordered):,}건",
        "# 생성: scripts/build_address_gazetteer.py",
    ]
    body = "\n".join(header + ordered) + "\n"
    args.out.write_bytes(gzip.compress(body.encode("utf-8")))
    print(f"[ok] {rows:,}행 → {len(ordered):,}건 → {args.out} "
          f"({args.out.stat().st_size:,} bytes)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
